import openpyxl
class HomePageData:

    test_HomePage_data= [{"FirstName":"Basava","Email": "Basavana003","Gender": "Male"}, {"FirstName":"Jadeyappa","Email":"Gouda", "Gender":"Male"}]

    @staticmethod
    def getTestData(test_case_name):
        dict = {}
        path = "C:\\Users\\Lenovo\\Desktop\\Revolution_2021\\123.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row = sheet.max_row
        print(row)
        column = sheet.max_column
        print(column)
        for r in range(1, row + 1):
            if sheet.cell(row=r, column=1).value ==test_case_name:
                for c in range(2, column + 1):
                    dict[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value
        return [dict]