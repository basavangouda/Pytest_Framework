import openpyxl

dict={}
path = "C:\\Users\\Lenovo\\Desktop\\Revolution_2021\\123.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
row = sheet.max_row
print(row)
column = sheet.max_column
print(column)
for r in range(1,row+1):
    if sheet.cell(row=r,column=1).value=='Test1':
        for c in range(2,column+1):
            dict[sheet.cell(row=1,column=c).value]=sheet.cell(row=r,column=c).value

print([dict])